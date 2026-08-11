import os
import sys
import tempfile
import openpyxl

# Add local path to import slm_document_parser
base_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, os.path.join(base_dir, "slm_document_parser"))

from slm_document_parser.document_parser import SLMDocumentParser

PASS = "\033[92mPASS\033[0m"
FAIL = "\033[91mFAIL\033[0m"

def run_test(name, fn):
    print(f"  ► {name}", end="... ", flush=True)
    try:
        msg = fn()
        print(f"[{PASS}] {msg or ''}", flush=True)
        return True
    except Exception as e:
        print(f"[{FAIL}] Error: {e}", flush=True)
        import traceback
        traceback.print_exc()
        return False

def test_docx_fallback():
    # Write a zip-structured mock docx
    import zipfile
    parser = SLMDocumentParser.__new__(SLMDocumentParser)
    
    with tempfile.NamedTemporaryFile(suffix=".docx", delete=False) as tmp:
        tmp_path = tmp.name
        
    try:
        with zipfile.ZipFile(tmp_path, 'w') as z:
            z.writestr('word/document.xml', '<?xml version="1.0"?><w:document><w:body><w:p><w:r><w:t>Hello DOCX Fallback World</w:t></w:r></w:p></w:body></w:document>')
            
        text = parser._extract_docx_fallback(tmp_path)
        assert "Hello DOCX Fallback World" in text, f"Got: {text}"
        return "Extracted DOCX XML zip contents successfully"
    finally:
        if os.path.exists(tmp_path):
            os.remove(tmp_path)

def test_pptx_fallback():
    import zipfile
    parser = SLMDocumentParser.__new__(SLMDocumentParser)
    
    with tempfile.NamedTemporaryFile(suffix=".pptx", delete=False) as tmp:
        tmp_path = tmp.name
        
    try:
        with zipfile.ZipFile(tmp_path, 'w') as z:
            z.writestr('ppt/slides/slide1.xml', '<?xml version="1.0"?><slide><a:t>Slide 1 content</a:t></slide>')
            z.writestr('ppt/slides/slide2.xml', '<?xml version="1.0"?><slide><a:t>Slide 2 text</a:t></slide>')
            
        slides = parser._extract_pptx_fallback(tmp_path)
        assert len(slides) == 2
        assert slides[0] == (1, "Slide 1 content")
        assert slides[1] == (2, "Slide 2 text")
        return "Extracted PPTX XML zip contents successfully"
    finally:
        if os.path.exists(tmp_path):
            os.remove(tmp_path)

def test_binary_string_extractor():
    parser = SLMDocumentParser.__new__(SLMDocumentParser)
    # Binary payload representing OLE text
    payload = b"\x00\x01\x02Hello ASCII Binary Payload\x00\x00\x00H\x00e\x00l\x00l\x00o\x00 \x00U\x00T\x00F\x001\x006\x00\x00\x00"
    text = parser._extract_strings_from_bytes(payload)
    assert "Hello ASCII Binary Payload" in text
    assert "Hello UTF16" in text
    return "Extracted ASCII and UTF-16LE characters from raw bytes successfully"

def test_chunk_linking():
    parser = SLMDocumentParser.__new__(SLMDocumentParser)
    # Create chunks to link
    chunks = [
        {
            "text": "Chunk A discusses AegisShield security modules.",
            "metadata": {
                "source": "doc1.pdf",
                "heading": "1. Introduction",
                "subheading": "1.1 Overview",
                "product": "AegisShield",
                "key_terms": ["security", "aegis", "intro"]
            }
        },
        {
            "text": "Chunk B details the overall Introduction scope.",
            "metadata": {
                "source": "doc1.pdf",
                "heading": "1. Introduction",
                "subheading": "1.2 Scope",
                "product": "",
                "key_terms": ["intro", "scope", "terms"]
            }
        },
        {
            "text": "Chunk C is unrelated, but mentions AegisShield.",
            "metadata": {
                "source": "doc1.pdf",
                "heading": "2. Advanced Setup",
                "subheading": "",
                "product": "AegisShield",
                "key_terms": ["setup", "aegis", "config"]
            }
        }
    ]
    
    linked = parser._link_chunks(chunks)
    # Chunk A (idx 0) should be linked to B (same heading '1. Introduction') and C (same product 'AegisShield')
    # and maybe shared terms
    assert 1 in linked[0]["metadata"]["related_chunks"]  # linked to index 1
    assert 2 in linked[0]["metadata"]["related_chunks"]  # linked to index 2
    
    # Chunk B (idx 1) linked to A (same heading)
    assert 0 in linked[1]["metadata"]["related_chunks"]
    
    # Chunk C (idx 2) linked to A (same product)
    assert 0 in linked[2]["metadata"]["related_chunks"]
    return "Linked section siblings and product references successfully"

def test_excel_export_and_append():
    parser = SLMDocumentParser.__new__(SLMDocumentParser)
    chunks1 = [
        {
            "text": "First chunk text.",
            "metadata": {
                "source": "document.pdf",
                "heading": "Header 1",
                "subheading": "Subheader 1",
                "product": "ProductX",
                "related_chunks": [1]
            }
        },
        {
            "text": "Second chunk text.",
            "metadata": {
                "source": "document.pdf",
                "heading": "Header 1",
                "subheading": "Subheader 2",
                "product": "ProductX",
                "related_chunks": [0]
            }
        }
    ]
    
    chunks2 = [
        {
            "text": "Third chunk text.",
            "metadata": {
                "source": "document2.pdf",
                "heading": "Header 2",
                "subheading": "",
                "product": "ProductY",
                "related_chunks": []
            }
        }
    ]
    
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        excel_path = tmp.name
        
    try:
        # Initial Export
        parser.export_chunks_to_excel(chunks1, excel_path, append=False)
        
        # Verify first export
        wb = openpyxl.load_workbook(excel_path)
        ws = wb.active
        assert ws.max_row == 3  # Header + 2 chunks
        assert ws.cell(row=2, column=1).value == 0  # index 0
        assert ws.cell(row=3, column=1).value == 1  # index 1
        assert ws.cell(row=2, column=3).value == "Header 1"
        assert ws.cell(row=2, column=5).value == "ProductX"
        assert ws.cell(row=2, column=6).value == "1"
        
        # Append Export
        parser.export_chunks_to_excel(chunks2, excel_path, append=True)
        
        # Verify append results
        wb = openpyxl.load_workbook(excel_path)
        ws = wb.active
        assert ws.max_row == 4  # Header + 3 chunks
        assert ws.cell(row=4, column=1).value == 2  # index incremented to 2
        assert ws.cell(row=4, column=2).value == "document2.pdf"
        assert ws.cell(row=4, column=3).value == "Header 2"
        assert ws.cell(row=4, column=5).value == "ProductY"
        assert ws.cell(row=4, column=7).value == "Third chunk text."
        
        return "Exported and appended chunks to Excel with index continuity successfully"
    finally:
        if os.path.exists(excel_path):
            os.remove(excel_path)

def test_semantic_chunker_streaming():
    parser = SLMDocumentParser.__new__(SLMDocumentParser)
    
    # Mock _segment_and_extract_chunks
    def mock_segment(full_markdown, source_name):
        return [
            {
                "text": "Chunk 1 content",
                "metadata": {
                    "source": os.path.basename(source_name),
                    "heading": "Heading 1",
                    "subheading": "",
                    "product": "Prod1",
                    "key_terms": ["term1"],
                    "format": "txt"
                }
            },
            {
                "text": "Chunk 2 content",
                "metadata": {
                    "source": os.path.basename(source_name),
                    "heading": "Heading 1",
                    "subheading": "",
                    "product": "Prod1",
                    "key_terms": ["term2"],
                    "format": "txt"
                }
            }
        ]
        
    parser._segment_and_extract_chunks = mock_segment
    
    with tempfile.NamedTemporaryFile(suffix=".txt", mode="w", delete=False) as f:
        f.write("Some dummy file text")
        tmp_path = f.name
        
    try:
        # Stream check
        stream = parser.parse_and_chunk_stream(tmp_path)
        chunks = list(stream)
        
        assert len(chunks) == 2
        assert chunks[0]["text"] == "Chunk 1 content"
        assert chunks[1]["text"] == "Chunk 2 content"
        assert chunks[0]["metadata"]["chunk_index"] == 0
        assert chunks[1]["metadata"]["chunk_index"] == 1
        # Linking check
        assert chunks[0]["metadata"]["related_chunks"] == [1]
        assert chunks[1]["metadata"]["related_chunks"] == [0]
        
        return "Streamed and linked semantic chunks page-by-page successfully"
    finally:
        if os.path.exists(tmp_path):
            os.remove(tmp_path)

def main():
    print("="*60)
    print("Running Visual Chunker & Exporter Test Suite")
    print("="*60)
    
    success = True
    success &= run_test("DOCX XML Fallback Parser", test_docx_fallback)
    success &= run_test("PPTX XML Fallback Parser", test_pptx_fallback)
    success &= run_test("OLE Binary String Extractor", test_binary_string_extractor)
    success &= run_test("Chunk Graph Linkage Builder", test_chunk_linking)
    success &= run_test("Excel Exporter and Appender", test_excel_export_and_append)
    success &= run_test("Semantic Chunker Stream Generator", test_semantic_chunker_streaming)
    
    print("="*60)
    if success:
        print("All visual chunker tests completed successfully!")
        sys.exit(0)
    else:
        print("Some visual chunker tests failed.")
        sys.exit(1)

if __name__ == "__main__":
    main()
